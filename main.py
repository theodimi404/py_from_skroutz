import bs4
import re
import requests
import sys
from urllib.parse import urlparse
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl.styles import Font

ascii_art = """
                 __                           ____   _                         _
 _ __   _   _   / _| _ __   ___   _ __ ___   / ___| | | __ _ __   ___   _   _ | |_  ____
| '_ \ | | | | | |_ | '__| / _ \ | '_ ` _ \  \___ \ | |/ /| '__| / _ \ | | | || __||_  /
| |_) || |_| | |  _|| |   | (_) || | | | | |  ___) ||   < | |   | (_) || |_| || |_  / /
| .__/  \__, | |_|  |_|    \___/ |_| |_| |_| |____/ |_|\_\|_|    \___/  \__,_| \__|/___|
|_|     |___/
"""

print(ascii_art)


def google_search(product):
    url = f'https://www.google.com/search?q=skroutz{product}'
    result = []
    res = requests.get(url)
    if res.status_code == 200:
        soup = bs4.BeautifulSoup(res.text, 'html.parser')
        for link in soup.find_all('a'):
            k = link.get('href')
            try:
                m = re.search("(?P<url>https?://[^\s]+)", k)
                n = m.group(0)
                rul = n.split('&')[0]
                if urlparse(rul).netloc == 'www.skroutz.gr':
                    result.append(rul)
                else:
                    continue
            except:
                continue
        if result:
            response = result[0]
        else:
            response = 'No song lyrics'
    else:
        response = 'Error 404'

    return response


product = input('Enter the product that you are looking: ')
url = google_search(product)

opts = Options()
opts.headless = True
browser = Chrome(options=opts)
browser.get(url)

html_source = browser.page_source

soup = bs4.BeautifulSoup(html_source, 'html.parser')
# print(soup)

prices = soup.findAll("a", {"class": "js-product-link product-link content-placeholder"})
names = soup.findAll("a", {"class": "js-product-link content-placeholder"})
description = soup.find("div", {"class": "simple-description js-description-html"})
rating = soup.find("div", {"class": "rating-average cf"})
number_of_users_rating = soup.find("div", {"class": "actual-rating "})
rating2 = soup.find("a", {"class": "rating big_stars"})

if not prices:
    print('This product does not exist or you need to provide more information')
    sys.exit()

# Saving data on Excel

wb = Workbook()

sheet = wb.create_sheet("Details", 0)

max_width = 0
for i in range(len(prices)):
    # Saving in columns 1 and 2, all the prices and the products' names
    sheet.cell(row=i + 2, column=1).value = prices[i].text
    sheet.cell(row=i + 2, column=2).value = names[i].text
    sheet.cell(row=i + 2, column=2).hyperlink = 'https://www.skroutz.gr' + names[i]['href']  # Hyperlinks regarding
    # the names
    # Adjusting the column B, which contains the names
    if len(names[i].text) > max_width:
        sheet.column_dimensions['B'].width = len(names[i].text)
        max_width = len(names[i].text)

# Saving the rating of the product
try:
    sheet.cell(row=2, column=3).value = rating2['title']
    sheet.column_dimensions['C'].width = len(rating2['title'])
except:
    sheet.cell(row=2, column=3).value = 'No ratings'

# Saving min and max prices, and defining titles
sheet.cell(row=4, column=3).value = 'Min Price'
sheet.cell(row=4, column=3).font = Font(bold=True)
sheet.cell(row=6, column=3).value = 'Max Price'
sheet.cell(row=6, column=3).font = Font(bold=True)
sheet.cell(row=5, column=3).value = sheet.cell(row=2, column=1).value
sheet.cell(row=7, column=3).value = sheet.cell(row=i + 2, column=1).value

sheet.cell(row=1, column=1).value = 'Prices'
sheet.cell(row=1, column=2).value = 'Names'
sheet.cell(row=1, column=3).value = 'Rating'

sheet.cell(row=1, column=1).font = Font(bold=True)
sheet.cell(row=1, column=2).font = Font(bold=True)
sheet.cell(row=1, column=3).font = Font(bold=True)

# Saving the SPECS
specs = soup.findAll("div", {"class": "spec-details"})
k = 1
max_i = 0
max_l = 0
for j in specs:
    title = j.find("h3", {"class": ""})  # Category name
    dt = j.findAll("dt", {"class": ""})  # Name of the SPEC
    span = j.findAll("span", {"class": ""})  # SPECS
    try:
        sheet.cell(row=k, column=4).value = title.text
        sheet.cell(row=k, column=4).font = Font(bold=True)
        k = k + 1  # All data are saved in one column, so we iterate through k rows
    except:
        print("")
    for i, l in zip(dt, span):
        try:
            # Saving name and spec
            sheet.cell(row=k, column=4).value = i.text
            sheet.cell(row=k, column=5).value = l.text
            k = k + 1
            if len(i.text) > max_i:
                sheet.column_dimensions['D'].width = len(i.text)
                max_i = len(i.text)
            if len(l.text) > max_l:
                sheet.column_dimensions['E'].width = len(l.text)
                max_l = len(l.text)
        except:
            print("")

wb.save(product + '.xlsx')
browser.quit()
